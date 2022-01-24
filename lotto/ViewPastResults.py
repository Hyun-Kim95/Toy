from openpyxl import load_workbook
import time
# 로또 과거번호 조회

def 과거등수(result, 전체숫자):
  max_value = [0,0]
  
  for chk, i in enumerate(전체숫자):
    noBonus = []
    for q in range(6):
      noBonus.append(i[q])
    cnt = 0
    for j in result:
      if j in noBonus:
        cnt += 1
    if cnt == 5:
      if i[-1] in result:
        cnt = 5.5
    if cnt > max_value[0]:
      max_value[0] = cnt
      max_value[1] = chk
  결과 = "과거기록 : "
  if max_value[0] == 6:
    결과 += "{}회차 1등".format(len(전체숫자) - max_value[1])
  elif max_value[0] == 5.5:
    결과 += "{}회차 2등".format(len(전체숫자) - max_value[1])
  elif max_value[0] == 5:
    결과 += "{}회차 3등".format(len(전체숫자) - max_value[1])
  elif max_value[0] == 4:
    결과 += "{}회차 4등".format(len(전체숫자) - max_value[1])
  elif max_value[0] == 3:
    결과 += "{}회차 5등".format(len(전체숫자) - max_value[1])
  else:
    결과 += "없음"
  
  return 결과

wb = load_workbook("C:\\Users\\User\\Desktop\\lotto\\excel.xlsx")
ws = wb.active

# 1회차 줄 확인을 위한 현재 회차 구함
a = ws.cell(row=4,column=2).value + 4
전체숫자 = []

for p in range(4,a):
  라인 = []
  for k in range(7):
    라인.append(ws.cell(row=p, column=14+k).value)
  전체숫자.append(라인)
wb.close()

while True:
  print("\n종료를 원하실 때는 'y'를 입력해주세요.")
  chk = input("\n조회하실 번호 6개를 띄어쓰기 한칸을 기준으로 입력해주세요 >> ").strip()
  
  if chk.strip() == 'y':
    print("5초 후에 종료합니다.")
    for i in range(5,0,-1):
      print(i)
      time.sleep(1)
    break
  
  chk = chk.split(" ")
  if len(chk) == 6:
    try:
      chk = [int(i) for i in chk]
      print(과거등수(chk, 전체숫자))
    except:
      print("\n숫자만 입력해주세요!")
  else:
    print("\n6개의 숫자를 입력해주세요!")