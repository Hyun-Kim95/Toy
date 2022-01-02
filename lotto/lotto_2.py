# 엑셀에서 최근 숫자와 다음에 올 숫자 등 분석
from openpyxl import load_workbook

wb = load_workbook("C:\\Hyun_Folder\\ETC\\lotto\\excel.xlsx")
ws = wb.active

# 1회차 줄 확인을 위한 현재 회차 구함
a = ws.cell(row=4,column=2).value + 4
최근숫자 = []
다음숫자 = []

# 라인 별 최근 수 구함
for num in range(6):
  최근숫자.append(ws.cell(row=4,column=14+num).value)
# 라인 별 다음 수 구함
i = 0
while(i < 6):
  라인 = []
  for j in range(5,a):
    if(ws.cell(row=j,column=14+i).value == 최근숫자[i]):
      if(ws.cell(row=j-1,column=14+i).value not in 라인):
        if(i > 0):
          if(ws.cell(row=j-1,column=14+i).value not in 다음숫자[i-1]):
            라인.append(ws.cell(row=j-1,column=14+i).value)
        else:
          라인.append(ws.cell(row=j-1,column=14+i).value)
    if(j == a-1):
      if(len(라인) == 0):
        최근숫자[i] -= 1
        i -= 1
      else:
        다음숫자.append(라인)
  i += 1

# 앞숫자 - 뒷숫자 (오차확인용)
cha = []
for number in range(6):
  yy = []
  for ten in range(10):
    yy.append(ws.cell(row=5+ten,column=14+number).value - ws.cell(row=6+ten,column=14+number).value)
  cha.append(yy)
wb.close()